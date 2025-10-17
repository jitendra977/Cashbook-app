from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from rest_framework.permissions import AllowAny
from datetime import datetime, timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import json
from rest_framework.decorators import action

from decimal import Decimal
from .models import TransactionType, TransactionCategory, Transaction, CashbookBalance
from store.models import StoreUser, Cashbook
from .serializers import (
    TransactionTypeSerializer,
    TransactionCategorySerializer,
    TransactionListSerializer,
    TransactionDetailSerializer,
    TransactionCreateUpdateSerializer,
    TransactionBulkCreateSerializer,
    TransactionSummarySerializer,
    CashbookBalanceSerializer
)


class TransactionTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing transaction types
    
    list: Get all transaction types
    retrieve: Get a specific transaction type
    create: Create a new transaction type
    update: Update a transaction type
    partial_update: Partially update a transaction type
    destroy: Delete a transaction type
    
    Custom actions:
    - active: Get only active transaction types
    - by_nature: Get transaction types filtered by nature (income/expense/transfer)
    """
    queryset = TransactionType.objects.all()
    serializer_class = TransactionTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter, DjangoFilterBackend]
    search_fields = ['name']
    ordering_fields = ['name', 'created_at']
    filterset_fields = ['nature', 'is_active']

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active transaction types"""
        queryset = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_nature(self, request):
        """Get transaction types by nature"""
        nature = request.query_params.get('nature')
        if not nature:
            return Response(
                {"error": "Nature parameter is required (income/expense/transfer)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if nature not in ['income', 'expense', 'transfer']:
            return Response(
                {"error": "Invalid nature. Must be income, expense, or transfer"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(nature=nature, is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TransactionCategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing transaction categories
    
    list: Get all transaction categories
    retrieve: Get a specific transaction category
    create: Create a new transaction category
    update: Update a transaction category
    partial_update: Partially update a transaction category
    destroy: Delete a transaction category
    
    Custom actions:
    - active: Get only active categories
    """
    queryset = TransactionCategory.objects.all()
    serializer_class = TransactionCategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    filterset_fields = ['is_active']

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active categories"""
        queryset = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TransactionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing transactions with comprehensive filtering and reporting
    
    Standard actions:
    - list: Get all transactions (filtered by user's store access)
    - retrieve: Get a specific transaction
    - create: Create a new transaction
    - update: Update a transaction
    - partial_update: Partially update a transaction
    - destroy: Delete a transaction
    
    Custom actions:
    - summary: Get transaction summary statistics
    - by_store: Get transactions for a specific store
    - by_cashbook: Get transactions for a specific cashbook
    - by_date_range: Get transactions within a date range
    - recent: Get recent transactions
    - pending: Get pending transactions
    - bulk_create: Create multiple transactions at once
    - export: Export transactions (returns data ready for CSV/Excel)
    """
    queryset = Transaction.objects.select_related(
        'cashbook', 'cashbook__store', 'type', 'category', 'created_by', 'updated_by'
    ).all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['cashbook', 'type', 'category', 'transaction_date', 'created_by','status']
    search_fields = ['description', 'reference_number', 'cashbook__name', 'created_by','transaction_id']
    ordering_fields = ['transaction_date', 'amount', 'created_at', 'created_by', 'updated_at']
    ordering = ['-transaction_date', '-created_at']

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action in ['create', 'update', 'partial_update']:
            return TransactionCreateUpdateSerializer
        elif self.action == 'list':
            return TransactionListSerializer
        elif self.action == 'bulk_create':
            return TransactionBulkCreateSerializer
        return TransactionDetailSerializer

    def get_queryset(self):
        """Filter transactions by user's store access"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # Get store IDs that the user has access to
        user_store_ids = StoreUser.objects.filter(user=user).values_list('store_id', flat=True)
        
        # Filter transactions by accessible stores
        return queryset.filter(cashbook__store_id__in=user_store_ids)

    def perform_create(self, serializer):
        """Set created_by when creating transaction"""
        serializer.save()

    def perform_update(self, serializer):
        """Set updated_by when updating transaction"""
        serializer.save()

    def perform_destroy(self, instance):
        """Reverse balance changes before deleting transaction"""
        cashbook = instance.cashbook
        amount = instance.amount
        nature = instance.type.nature
        
        # Reverse the balance effect
        if nature == 'income':
            cashbook.current_balance -= amount
        elif nature == 'expense':
            cashbook.current_balance += amount
        
        cashbook.save(update_fields=['current_balance', 'updated_at'])
        
        # Delete the transaction
        instance.delete()

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Get transaction summary with totals and statistics
        Query params: store, cashbook, start_date, end_date
        """
        queryset = self._apply_filters(request, self.get_queryset())
        
        # Calculate summary statistics
        summary_data = queryset.aggregate(
            total_transactions=Count('id'),
            total_amount=Sum('amount'),
            total_income=Sum('amount', filter=Q(type__nature='income')),
            total_expense=Sum('amount', filter=Q(type__nature='expense')),
            average_transaction=Avg('amount')
        )
        
        # Handle None values
        for key in ['total_amount', 'total_income', 'total_expense', 'average_transaction']:
            if summary_data[key] is None:
                summary_data[key] = 0
        
        # Calculate net balance
        summary_data['net_balance'] = (summary_data['total_income'] or 0) - (summary_data['total_expense'] or 0)
        
        # Add date range info
        summary_data['period_start'] = request.query_params.get('start_date')
        summary_data['period_end'] = request.query_params.get('end_date')
        
        serializer = TransactionSummarySerializer(summary_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_store(self, request):
        """Get transactions for a specific store"""
        store_id = request.query_params.get('store')
        if not store_id:
            return Response(
                {"error": "Store ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify user has access to this store
        if not self._verify_store_access(request.user, store_id):
            return Response(
                {"error": "You don't have access to this store"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        queryset = self.get_queryset().filter(cashbook__store_id=store_id)
        queryset = self._apply_filters(request, queryset)
        
        return self._paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def by_cashbook(self, request):
        """Get transactions for a specific cashbook"""
        cashbook_id = request.query_params.get('cashbook')
        if not cashbook_id:
            return Response(
                {"error": "Cashbook ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(cashbook_id=cashbook_id)
        
        # Verify access through the queryset (already filtered by user's stores)
        if not queryset.exists():
            try:
                cashbook = Cashbook.objects.get(id=cashbook_id)
                if not self._verify_store_access(request.user, cashbook.store_id):
                    return Response(
                        {"error": "You don't have access to this cashbook"},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except Cashbook.DoesNotExist:
                return Response(
                    {"error": "Cashbook not found"},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        queryset = self._apply_filters(request, queryset)
        return self._paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def by_date_range(self, request):
        """Get transactions within a date range"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date or not end_date:
            return Response(
                {"error": "Both start_date and end_date are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            queryset = self.get_queryset().filter(
                transaction_date__gte=start_date,
                transaction_date__lte=end_date
            )
            queryset = self.filter_queryset(queryset)
            return self._paginated_response(queryset)
        except Exception as e:
            return Response(
                {"error": f"Invalid date format: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent transactions (last 7 days by default)"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        queryset = self.get_queryset().filter(transaction_date__gte=start_date)
        queryset = self._apply_filters(request, queryset)
        
        return self._paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending transactions"""
        queryset = self.get_queryset().filter(status='pending')
        queryset = self._apply_filters(request, queryset)
        
        return self._paginated_response(queryset)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple transactions at once"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            transactions = serializer.save()
            response_serializer = TransactionDetailSerializer(transactions, many=True)
            return Response(
                response_serializer.data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to create transactions: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def export(self, request):
        """
        Export transactions in multiple formats: Excel, JSON, or PDF
        Query params:
        - format: 'excel', 'json', or 'pdf' (default: 'json')
        - All standard filters: store, cashbook, start_date, end_date, type, category, status
        """
        try:
            # Apply all filters from request
            queryset = self._apply_filters(request, self.get_queryset())
            
            # Optional: Limit export size to prevent performance issues
            max_export_size = 10000
            total_count = queryset.count()
            
            if total_count > max_export_size:
                return Response(
                    {
                        "error": f"Cannot export more than {max_export_size} records. "
                                f"Your query returned {total_count} records. "
                                f"Please apply additional filters to reduce the result set."
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # If no results, return early with message
            if total_count == 0:
                return Response({
                    "count": 0,
                    "data": [],
                    "message": "No transactions found matching the criteria"
                })
            
            # Get export format from query params
            export_format = request.query_params.get('format', 'json').lower()
            
            # Validate format
            if export_format not in ['excel', 'json', 'pdf']:
                return Response(
                    {"error": "Invalid format. Must be 'excel', 'json', or 'pdf'"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get filter info for metadata
            filters_applied = self._get_applied_filters(request)
            
            # Route to appropriate export method
            if export_format == 'excel':
                return self._export_excel(queryset, filters_applied, request.user)
            elif export_format == 'pdf':
                return self._export_pdf(queryset, filters_applied, request.user)
            else:  # json
                return self._export_json(queryset, filters_applied, request.user)
                
        except ValueError as e:
            return Response(
                {"error": f"Invalid parameter value: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"Export failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _export_json(self, queryset, filters_applied, user):
        """Export transactions as JSON"""
        serializer = TransactionListSerializer(queryset, many=True)
        
        response_data = {
            "success": True,
            "count": queryset.count(),
            "data": serializer.data,
            "exported_at": timezone.now().isoformat(),
            "exported_by": user.username,
            "filters_applied": filters_applied,
        }
        
        return Response(response_data)
    
    def _export_excel(self, queryset, filters_applied, user):
        """Export transactions as Excel file"""
        # Create workbook in memory
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:J1')
        ws['A1'] = "Transaction Export Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Add metadata
        row = 3
        metadata = [
            ("Exported By:", user.username),
            ("Exported At:", timezone.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Records:", str(queryset.count()))
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Add filters if applied
        if filters_applied:
            row += 1
            ws[f'A{row}'] = "Filters Applied:"
            ws[f'A{row}'].font = Font(bold=True)
            for key, value in filters_applied.items():
                row += 1
                ws[f'A{row}'] = f"  {key.replace('_', ' ').title()}:"
                ws[f'B{row}'] = str(value)
        
        # Add header row
        row += 2
        headers = [
            'ID', 'Date', 'Cashbook', 'Store', 'Type', 
            'Category', 'Amount', 'Description', 'Status', 'Reference'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data rows
        for transaction in queryset:
            row += 1
            data = [
                transaction.id,
                transaction.transaction_date.strftime("%Y-%m-%d") if transaction.transaction_date else '',
                transaction.cashbook.name if transaction.cashbook else '',
                transaction.cashbook.store.name if transaction.cashbook and transaction.cashbook.store else '',
                transaction.type.name if transaction.type else '',
                transaction.category.name if transaction.category else '',
                float(transaction.amount) if transaction.amount else 0.0,
                transaction.description or '',
                transaction.status or '',
                transaction.reference_number or ''
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = thin_border
                
                # Format amount column
                if col == 7:  # Amount column
                    cell.number_format = '#,##0.00'
        
        # Adjust column widths
        column_widths = [12, 12, 20, 20, 15, 15, 12, 30, 12, 15]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Save workbook to bytes buffer
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'transactions_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _export_pdf(self, queryset, filters_applied, user):
        """Export transactions as PDF file"""
        # Create bytes buffer for PDF
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch)
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2E5BFF'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        # Add title
        title = Paragraph("Transaction Export Report", title_style)
        elements.append(title)
        
        # Add metadata
        metadata_text = f"""
        <b>Exported By:</b> {user.username}<br/>
        <b>Exported At:</b> {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}<br/>
        <b>Total Records:</b> {queryset.count()}<br/>
        """
        
        if filters_applied:
            metadata_text += "<b>Filters Applied:</b><br/>"
            for key, value in filters_applied.items():
                metadata_text += f"&nbsp;&nbsp;{key.replace('_', ' ').title()}: {value}<br/>"
        
        metadata = Paragraph(metadata_text, styles["Normal"])
        elements.append(metadata)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare table data
        table_data = [['ID', 'Date', 'Cashbook', 'Type', 'Category', 'Amount', 'Status']]
        
        for transaction in queryset:
            table_data.append([
                str(transaction.id),
                transaction.transaction_date.strftime("%Y-%m-%d") if transaction.transaction_date else '',
                (transaction.cashbook.name[:20] + '...') if transaction.cashbook and len(transaction.cashbook.name) > 20 else (transaction.cashbook.name if transaction.cashbook else ''),
                transaction.type.name if transaction.type else '',
                transaction.category.name if transaction.category else '',
                f"${float(transaction.amount):,.2f}" if transaction.amount else '$0.00',
                transaction.status or ''
            ])
        
        # Create table
        col_widths = [0.8*inch, 1*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1*inch, 0.8*inch]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Style the table
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5BFF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data style
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),  # Amount column
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/pdf'
        )
        filename = f'transactions_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    def _get_applied_filters(self, request):
        """Get dictionary of applied filters"""
        filters_applied = {}
        filter_params = ['store', 'cashbook', 'start_date', 'end_date', 'type', 'category', 'status']
        
        for param in filter_params:
            value = request.query_params.get(param)
            if value:
                filters_applied[param] = value
        
        return filters_applied
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Get transactions grouped by category with totals"""
        queryset = self._apply_filters(request, self.get_queryset())
        
        category_summary = queryset.values(
            'category__name', 'category__id'
        ).annotate(
            total_amount=Sum('amount'),
            transaction_count=Count('id')
        ).order_by('-total_amount')
        
        return Response(list(category_summary))

    @action(detail=False, methods=['get'])
    def by_type(self, request):
        """Get transactions grouped by type with totals"""
        queryset = self._apply_filters(request, self.get_queryset())
        
        type_summary = queryset.values(
            'type__name', 'type__id', 'type__nature'
        ).annotate(
            total_amount=Sum('amount'),
            transaction_count=Count('id')
        ).order_by('-total_amount')
        
        return Response(list(type_summary))

    # Helper methods
    def _apply_filters(self, request, queryset):
        """Apply common filters from query params"""
        # Store filter
        store_id = request.query_params.get('store')
        if store_id:
            queryset = queryset.filter(cashbook__store_id=store_id)
        
        # Cashbook filter
        cashbook_id = request.query_params.get('cashbook')
        if cashbook_id:
            queryset = queryset.filter(cashbook_id=cashbook_id)
        
        # Date filters
        start_date = request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(transaction_date__lte=end_date)
        
        # Apply standard filters
        queryset = self.filter_queryset(queryset)
        
        return queryset

    def _verify_store_access(self, user, store_id):
        """Verify user has access to a store"""
        return StoreUser.objects.filter(user=user, store_id=store_id).exists()

    def _paginated_response(self, queryset):
        """Return paginated response"""
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class CashbookBalanceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing cashbook balances (read-only)
    
    list: Get all cashbook balances
    retrieve: Get a specific cashbook balance
    
    Custom actions:
    - by_cashbook: Get balances for a specific cashbook
    - by_date_range: Get balances within a date range
    - latest: Get the latest balance for each cashbook
    """
    queryset = CashbookBalance.objects.select_related('cashbook', 'cashbook__store').all()
    serializer_class = CashbookBalanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['cashbook', 'date']
    ordering_fields = ['date']
    ordering = ['-date']

    def get_queryset(self):
        """Filter by user's store access"""
        queryset = super().get_queryset()
        user = self.request.user
        
        user_store_ids = StoreUser.objects.filter(user=user).values_list('store_id', flat=True)
        return queryset.filter(cashbook__store_id__in=user_store_ids)

    @action(detail=False, methods=['get'])
    def by_cashbook(self, request):
        """Get balances for a specific cashbook"""
        cashbook_id = request.query_params.get('cashbook')
        if not cashbook_id:
            return Response(
                {"error": "Cashbook ID is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(cashbook_id=cashbook_id)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get the latest balance for each cashbook"""
        from django.db.models import Max
        
        queryset = self.get_queryset()
        
        # Get latest date for each cashbook
        latest_dates = queryset.values('cashbook').annotate(
            latest_date=Max('date')
        )
        
        # Build Q objects for filtering
        q_objects = Q()
        for item in latest_dates:
            q_objects |= Q(cashbook=item['cashbook'], date=item['latest_date'])
        
        latest_balances = queryset.filter(q_objects)
        serializer = self.get_serializer(latest_balances, many=True)
        return Response(serializer.data)